from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from serial.tools import list_ports
import json
import apsw
import os
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import openpyxl

alphabet = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"]

class checkBack(APIView):
    def get(self, request):
        return Response()

class getInts(APIView):
    def get(self, request):
        ports = list_ports.comports()
        portsDevice = []
        for port in ports:
            portsDevice.append(port.device)
        return Response({
            "ints": portsDevice
        })

def getData(file):
    try: #log 2400
        with apsw.Connection(':memory:') as db:
            db.deserialize('main', file)
            cursor = db.cursor()
            cursor.execute('SELECT * FROM Frames;')
            frames = cursor.fetchall()
            log_data = []
            for frame in frames:
                log_data.append({
                    "id": frame[0],
                    "freq_arr": frame[1],
                    "norm_freq_arr": frame[2],
                    "time": frame[3],
                    "zone_state": frame[4],
                    "crit_level": frame[5],
                    "note": frame[6]
                })
            return {
                "result": True,
                "log_data": log_data,
                "type": 2.4
            }
    except:
        try: #log 915
            with apsw.Connection(':memory:') as db:
                db.deserialize('main', file)
                cursor = db.cursor()
                cursor.execute('SELECT * FROM Frames;')
                frames = cursor.fetchall()
                log_data = []
                try:
                    for frame in frames:
                        log_data.append({
                            "id": frame[0],
                            "freq_arr": frame[1],
                            "time": frame[2],
                            "note": frame[3],
                            "zone_state": frame[4]
                        })
                except:
                    for frame in frames:
                        log_data.append({
                            "id": frame[0],
                            "freq_arr": frame[1],
                            "time": frame[2],
                            "note": frame[3]
                        })
                return {
                    "result": True,
                    "log_data": log_data,
                    "type": 443
                }
        except: 
            try: #log Multi
                with apsw.Connection(':memory:') as db:
                    db.deserialize('main', file)
                    cursor = db.cursor()
                    cursor.execute('SELECT * FROM FramesMulti;')
                    frames = cursor.fetchall()
                    log_data = []
                    for frame in frames:
                        log_data.append({
                            "id": frame[0],
                            "freq_arr": frame[1],
                            "time": frame[2],
                            "note": frame[3]
                        })
                    return {
                        "result": True,
                        "log_data": log_data,
                        "type": "multi"
                    }
            except:
                return {
                    "result": False
                }

class getLogData(APIView):
    def post(self, request):
        file = request.data['file'].read()
        return Response(getData(file))
        
            
class getLogDataMult(APIView):
    def post(self, request):
        file_1 = request.data['file_1'].read()
        file_2 = request.data["file_2"].read()
        res_1 = getData(file_1)
        res_2 = getData(file_2)
        if res_1["result"] == False or res_2["result"] == False:
            return Response({
                "result": False 
            })
        else:
            if res_1["type"] == "multi" or res_2["type"] == "multi":
                return Response({
                    "result": False
                })
            else:
                return Response({
                    "result": True,
                    "res_1": res_1,
                    "res_2": res_2
                })


def writeDataXLS(data, file_name):
    wb = Workbook()
    ws = wb.active
    rows = [["time", "width", "", "time", "height", "", "time", "zone_state"]]
    for i in range(len(data["time"])):
        rows.append([
            data["time"][i], data["width"][i], "", data["time"][i], data["height"][i], "", data["time"][i], data["zone_state"][i]
        ])
    
    for row in rows:
        ws.append(row)

    #горизонтальная покраска
    yellow_color = openpyxl.styles.colors.Color(rgb='FFEB31')
    yelow_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=yellow_color)
    for i in range(1, 3):
        ws[f"{alphabet[i-1]}1"].fill = yelow_fill
    for i in range(4, 6):
        ws[f"{alphabet[i-1]}1"].fill = yelow_fill
    for i in range(7, 9):
        ws[f"{alphabet[i-1]}1"].fill = yelow_fill
    
    #Вертикальная покраска
    green_color = openpyxl.styles.colors.Color(rgb='42FF16')
    green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green_color)
    for i in range(2, len(data["time"])+2):
        ws[f"A{i}"].fill = green_fill
        ws[f"D{i}"].fill = green_fill
        ws[f"G{i}"].fill = green_fill

    #основные графики
    chart1 = ScatterChart()
    chart1.height = 14
    chart1.width = 24
    chart1.title = "Graph of peak width changes depending on time"
    chart1.style = 13
    chart1.x_axis.title = 'Time'
    chart1.y_axis.title = 'Width'
    xvalues1 = Reference(ws, min_col=1, min_row=2, max_row=len(data["time"])+1)
    yvalues1 = Reference(ws, min_col=2, min_row=2, max_row=len(data["time"])+1)
    series1 = Series(yvalues1, xvalues1, title_from_data=False)
    chart1.series.append(series1)
    ws.add_chart(chart1, "J2")

    chart2 = ScatterChart()
    chart2.height = 14
    chart2.width = 24
    chart2.title = "Graph of peak height changes depending on time"
    chart2.style = 15
    chart2.x_axis.title = 'Time'
    chart2.y_axis.title = 'Height'
    xvalues2 = Reference(ws, min_col=4, min_row=2, max_row=len(data["time"])+1)
    yvalues2 = Reference(ws, min_col=5, min_row=2, max_row=len(data["time"])+1)
    series2 = Series(yvalues2, xvalues2, title_from_data=False)
    chart2.series.append(series2)
    ws.add_chart(chart2, "J29")

    chart3 = ScatterChart()
    chart3.height = 14
    chart3.width = 24
    chart3.title = "Presence(1)/absence(0) of UAVs"
    chart3.style = 16
    chart3.x_axis.title = 'Time'
    chart3.y_axis.title = 'Value'
    xvalues3 = Reference(ws, min_col=7, min_row=2, max_row=len(data["time"])+1)
    yvalues3 = Reference(ws, min_col=8, min_row=2, max_row=len(data["time"])+1)
    series3 = Series(yvalues3, xvalues3, title_from_data=False)
    chart3.series.append(series3)
    ws.add_chart(chart3, "W2")

    wb.save(f"{os.getcwd().replace('fullstack_django', 'reports')}/{file_name[:len(file_name)-3]}.xlsx")

def writeGraph(x_data, y_data, style, start, ws, x_labels, y_labels, title, start_col):
    chart = ScatterChart()
    chart.height = 14
    chart.width = 24
    chart.title = title
    chart.style = style
    chart.x_axis.title = x_labels
    chart.y_axis.title = y_labels
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=len(x_data)+1)
    yvalues = Reference(ws, min_col=start_col, min_row=2, max_row=len(y_data)+1)
    series = Series(yvalues, xvalues, title_from_data=False)
    chart.series.append(series)
    ws.add_chart(chart, start)

def writeDataXLS915(data, file_name):
    wb = Workbook()
    ws = wb.active
    rows = [["time", "active", "", "max", "", "av_rssi", "", f"under_{data['rssi_level']}"]]
    for i in range(len(data["time_data"])):
        rows.append([data["time_data"][i], data["activity_data"][i], "", data["max_data"][i], "", data["av_rssi_data"][i], "", data["under_rssi_data"][i]])
    for row in rows:
        ws.append(row) 

    #горизонтальная покраска
    yellow_color = openpyxl.styles.colors.Color(rgb='FFEB31')
    yelow_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=yellow_color)
    for i in range(1, 3):
        ws[f"{alphabet[i-1]}1"].fill = yelow_fill
    ws["D1"].fill = yelow_fill
    ws["F1"].fill = yelow_fill
    ws["H1"].fill = yelow_fill
    
    #Вертикальная покраска
    green_color = openpyxl.styles.colors.Color(rgb='42FF16')
    green_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=green_color)
    for i in range(2, len(data["time_data"])+2):
        ws[f"A{i}"].fill = green_fill

    writeGraph(data["time_data"], data["activity_data"], 13, "j2", ws, "Time", "Activity", "Graph of changes in the workload of the activity", 2)
    writeGraph(data["time_data"], data["max_data"], 15, "j29", ws, "Time", "Max_value", "Graph of changes in the workload of the max value", 4)
    writeGraph(data["time_data"], data["av_rssi_data"], 18, "W2", ws, "Time", "av_rssi", "Graph of changes in the workload of the rssi", 6)
    writeGraph(data["time_data"], data["under_rssi_data"], 13, "W29", ws, "Time", f"under_{data['rssi_level']}", "Graph of changes in the workload under rssi", 8)


    wb.save(f"{os.getcwd().replace('fullstack_django', 'reports')}/{file_name[:len(file_name)-3]}.xlsx")

class analizeLog(APIView):
    def post(self, request):
        data = json.loads(request.data["data"])
        file_name = request.data["file_name"]
        writeDataXLS(data, file_name)
        return Response({})

class analizeLog915(APIView):
    def post(self, request):
        data = json.loads(request.data["data"])
        file_name = request.data["file_name"]
        writeDataXLS915(data, file_name)
        return Response({})
